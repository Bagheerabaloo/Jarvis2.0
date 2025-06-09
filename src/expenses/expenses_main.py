import os
import shutil
from dataclasses import dataclass
from enum import Enum
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment

import pandas as pd
from pandas_ta import df_month_to_date, df_dates

EXPENSES_FOLDER = r"C:\Users\Vale\Documents\Spese"


class TransactionFields(Enum):
    ACCOUNTING_DATE = "Data Contabile"
    VALUE_DATE = "Data Valuta"
    TRANSACTION = "Operazione"
    AMOUNT = "Importo"
    CURRENCY = "Valuta"
    DETAILS = "Dettagli"
    ACCOUNT = "Conto o carta"
    CATEGORY = "Categoria"
    MACRO_CATEGORY = "Macro Categoria"
    GIFT = "Regalo"
    CURRENCY_CREDITS = "Accrediti in valuta"
    CURRENCY_DEBITS = "Addebiti in valuta"
    CREDIT_CARD = "Credit Card"

    @classmethod
    def get_common_columns(cls):
        """Return a list of values for the selected common columns."""
        common_fields = [
            cls.ACCOUNTING_DATE,
            cls.VALUE_DATE,
            cls.TRANSACTION,
            cls.AMOUNT,
            cls.CURRENCY,
            cls.DETAILS,
            cls.ACCOUNT,
            cls.CATEGORY,
            cls.MACRO_CATEGORY,
            cls.GIFT,
            cls.CURRENCY_CREDITS,
            cls.CURRENCY_DEBITS,
            cls.CREDIT_CARD,
        ]
        return [field.value for field in common_fields]


@dataclass
class Expenses:
    back_up_done: bool = False

    def main(self):
        # __ read files __
        old_txs = self.read_old_txs()  # old txs
        new_txs = self.read_new_txs()  # new txs
        cc_details = self.read_cc_details()  # cc details

        # __ if there are any missing Data Contabile in the old txs, try to fill them with cc_details __
        result = self.fill_missing_data_contabile(old_txs, cc_details)

        if result:
            old_txs = self.read_old_txs()

        # __ filter out txs already present in the old txs DataFrame __
        new_txs_filt = self.filter_new_txs(old_txs, new_txs)

        # __ add cc details to new txs __
        new_txs_filt_cc = self.add_cc_details(new_txs_filt, cc_details)

        # __ assign categories and macro categories __
        new_txs_filt_cc_cats = self.add_categories(old_txs, new_txs_filt_cc)
        new_txs_filt_cc_cats["Regalo"] = ""

        # __ filter columns __
        common_cols = TransactionFields.get_common_columns()
        remaining_cols = [x for x in new_txs_filt_cc_cats.columns if x not in common_cols]
        result = new_txs_filt_cc_cats[common_cols + remaining_cols]
        print([x for x in old_txs.columns if x not in result.columns])
        print([x for x in result.columns if x not in old_txs.columns])
        assert len([x for x in result.columns if x not in old_txs.columns]) == 0
        result = result[[x for x in old_txs.columns if x in result.columns]]

        # __ back_up old txs: rename and move old file to back-up folder __
        if not self.back_up_done:
            self.back_up_old_txs()

        # __ update old txs with new txs __
        self.update_txs(old_txs, result)

        # __ move files to old folder __
        self.move_files_to_old_folder()

        print('end')

    @staticmethod
    def read_old_txs() -> pd.DataFrame:
        txs_path = fr"{EXPENSES_FOLDER}\txs.xlsx"
        txs_df = pd.read_excel(txs_path)
        txs_df["Data Contabile"] = pd.to_datetime(txs_df["Data Contabile"]).dt.date
        txs_df["Data Valuta"] = pd.to_datetime(txs_df["Data Valuta"]).dt.date
        txs_df = txs_df.sort_values(by='Data Valuta')
        return txs_df

    @staticmethod
    def read_new_txs() -> pd.DataFrame:
        new_txs_files = [x for x in os.listdir(EXPENSES_FOLDER) if x.startswith('lista_completa') and x.endswith('.xlsx') and x != 'txs.xlsx']
        new_txs_df_parts = []
        for file in new_txs_files:
            # __ read new txs starting from row 21 __
            new_txs_df = pd.read_excel(fr"{EXPENSES_FOLDER}\{file}", skiprows=20)
            new_txs_df_parts.append(new_txs_df)
        new_txs_df = pd.concat(new_txs_df_parts)
        new_txs_df = new_txs_df.drop_duplicates()
        new_txs_df = new_txs_df.sort_values(by='Data')

        def is_credit_card(x):
            return "Y" if 'CLASSIC CARD VISA' in x else "N"

        new_txs_df["Credit Card"] = new_txs_df["Conto o carta"].apply(is_credit_card)
        new_txs_df = new_txs_df.rename(columns={'Data': 'Data Valuta'})
        new_txs_df["Data Valuta"] = pd.to_datetime(new_txs_df["Data Valuta"]).dt.date

        return new_txs_df

    @staticmethod
    def read_cc_details() -> pd.DataFrame:
        cc_details_files = [x for x in os.listdir(EXPENSES_FOLDER) if x.startswith('Movimenti_Carta') and x.endswith('.xlsx') and x != 'txs.xlsx']
        cc_details_df_parts = []
        for file in cc_details_files:
            cc_details_df = pd.read_excel(fr"{EXPENSES_FOLDER}\{file}", skiprows=29)
            cc_details_df_parts.append(cc_details_df)
        cc_details_df = pd.concat(cc_details_df_parts)
        cc_details_df = cc_details_df.drop_duplicates()

        cc_details_df = cc_details_df.dropna(subset=['Data valuta'])
        cc_details_df = cc_details_df.rename(columns={'Data valuta': 'Data Valuta', 'Descrizione': 'Operazione'})
        cc_details_df["Data Valuta"] = pd.to_datetime(cc_details_df["Data Valuta"]).dt.date
        cc_details_df["Data contabile"] = pd.to_datetime(cc_details_df["Data contabile"]).dt.date
        cc_details_df["Accrediti"] = cc_details_df["Accrediti"].fillna(0.0)
        cc_details_df["Addebiti"] = cc_details_df["Addebiti"].fillna(0.0)
        cc_details_df["Importo"] = cc_details_df["Accrediti"] - cc_details_df["Addebiti"]
        cc_details_df = cc_details_df.drop(columns=["Accrediti", "Addebiti"])
        cc_details_df = cc_details_df.rename(columns={"Data contabile": "Data Contabile"})
        cc_details_df = cc_details_df.sort_values(by='Data Valuta')

        sorted_cols = ["Data Valuta", "Operazione", "Importo"]
        remaining_cols = [x for x in cc_details_df.columns if x not in sorted_cols]
        cc_details_df = cc_details_df[sorted_cols + remaining_cols]

        return cc_details_df

    def fill_missing_data_contabile(self, old_txs: pd.DataFrame, cc_details: pd.DataFrame) -> bool:
        """If there are any missing Data Contabile in the old txs, try to fill them with cc_details.
        Return True if any missing Data Contabile was filled, False otherwise"""
        if old_txs["Data Contabile"].isnull().any():
            filled_txs = self.get_empty_data_contabile(old_txs, cc_details)
            if len(filled_txs[filled_txs["Data Contabile"].notna()]) > 0:
                self.back_up_old_txs()
                self.update_data_contabile(filled_txs)
                return True

            if filled_txs["Data Contabile"].isnull().any():
                raise ValueError("Missing Data Contabile in old txs tx. Please remove manually the rows with missing Data Contabile.")
        return False

    def get_empty_data_contabile(self, txs_old: pd.DataFrame, cc_details: pd.DataFrame):
        txs = txs_old[txs_old["Data Contabile"].notna()]
        empty_txs = txs_old[txs_old["Data Contabile"].isnull()]
        assert len(empty_txs) + len(txs) == len(txs_old)

        empty_txs = empty_txs.drop(columns=["Data Contabile", "Operazione_cc", "Addebiti in valuta", "Accrediti in valuta"])
        empty_txs = empty_txs.reset_index()
        filled_txs = self.add_cc_details(empty_txs, cc_details)

        return filled_txs

    def filter_new_txs(self, txs_old: pd.DataFrame, txs_new: pd.DataFrame):
        # __ find perfect matches __
        subset_columns = ["Data Valuta", "Operazione", "Importo", "Data Contabile"]
        joined_df = txs_new.merge(txs_old[subset_columns], on=["Data Valuta", "Operazione", "Importo"], how="left")
        joined_df["perfect_match"] = joined_df["Data Contabile"].notna()
        joined_df = joined_df[list(txs_new.columns) + ["perfect_match"]]

        # __ find weaker matches __
        joined_df = joined_df.merge(
            txs_old[subset_columns].rename(columns={"Operazione": "Operazione_old"}),
            on=["Data Valuta", "Importo"],
            how="left")
        joined_df["weaker_match"] = joined_df["Data Contabile"].notna()
        joined_df = joined_df[list(txs_new.columns) + ["Operazione_old", "perfect_match", "weaker_match"]]
        joined_df["Operazione_old"] = joined_df["Operazione_old"].fillna("")

        def is_similar_description(x):
            if x["Operazione_old"] == "":
                return False
            if (x["Operazione"].lower() in x["Operazione_old"].lower()
                    or x["Operazione_old"].lower() in x["Operazione"].lower()):
                return True
            elif x['Operazione'].split(" ")[0].strip("*").lower() == x['Operazione_old'].split(" ")[0].strip("*").lower():
                return True
            return False

        joined_df["similar_description"] = joined_df.apply(is_similar_description, axis=1)
        sorted_cols = ["Data Valuta", "Operazione", "Importo", "Operazione_old", "perfect_match", "weaker_match", "similar_description"]
        remaining_cols = [x for x in joined_df.columns if x not in sorted_cols]
        joined_df = joined_df[sorted_cols + remaining_cols]

        # __ print new txs that have a weaker match and have similar description __
        bool_c = (joined_df["weaker_match"]) & (joined_df["similar_description"])
        to_remove = joined_df[bool_c].reset_index(drop=True)
        for idx, row in to_remove.iterrows():
            print("SKIPPED ", row["Data Valuta"], row["Importo"], row["Operazione"].rjust(25), row["Operazione_old"])

        # remove weaker matches with similar description
        txs_new_filtered = joined_df[~bool_c].reset_index(drop=True)

        return txs_new_filtered

    def add_cc_details(self, txs_new: pd.DataFrame, cc_details: pd.DataFrame):
        # __ find matches __
        joined_df = txs_new.merge(
            cc_details.rename(columns={"Operazione": "Operazione_cc"}),
            on=["Data Valuta", "Importo"],
            how="left")
        joined_df["match"] = joined_df["Data Contabile"].notna()
        joined_df["Operazione_cc"] = joined_df["Operazione_cc"].fillna("")

        def is_similar_description(x):
            if x["Operazione_cc"] == "":
                return False
            if (x["Operazione"].lower() in x["Operazione_cc"].lower()
                    or x["Operazione_cc"].lower() in x["Operazione"].lower()):
                return True
            elif x['Operazione'].split(" ")[0].strip("*").lower() == x['Operazione_cc'].split(" ")[0].strip("*").lower():
                return True
            elif x['Operazione'][5].lower() == x['Operazione_cc'][5].lower():
                return True
            return False

        joined_df["cc_similar_description"] = joined_df.apply(is_similar_description, axis=1)
        sorted_cols = ["Data Valuta", "Operazione", "Importo", "Credit Card", "Operazione_cc", "match", "cc_similar_description"]
        remaining_cols = [x for x in joined_df.columns if x not in sorted_cols]
        joined_df = joined_df[sorted_cols + remaining_cols]

        # __ print new txs of credit card that don't have a match and a similar description __
        bool_c = (joined_df["Credit Card"] == "Y") & (~joined_df["match"]) & (~joined_df["cc_similar_description"])
        to_remove = joined_df[bool_c].reset_index(drop=True)
        for idx, row in to_remove.iterrows():
            print("CREDIT CARD NO DETAILS: ", row["Data Valuta"], row["Importo"], row["Operazione"].rjust(25), row["Operazione_cc"])

        # __ only for txs where Credit Card = Y, set Data Contabile = Data Valuta __
        bool_c = (joined_df["Credit Card"] != "Y")
        joined_df.loc[bool_c, "Data Contabile"] = joined_df.loc[bool_c, "Data Valuta"]
        try:
            joined_df = joined_df.rename(columns={"Categoria ": "Categoria"})
        except KeyError:
            pass
        sorted_col = [
            "Data Contabile", "Data Valuta", "Operazione", "Valuta", "Dettagli", "Conto o carta",
            "Categoria", "Accrediti in valuta", "Addebiti in valuta", "Credit Card"]
        remaining_cols = [x for x in joined_df.columns if x not in sorted_col]
        joined_df = joined_df[sorted_col + remaining_cols]

        return joined_df

    def add_categories(self, txs_old: pd.DataFrame, txs_new: pd.DataFrame):

        mapping_df = pd.read_csv(fr"{EXPENSES_FOLDER}\mapping.csv")
        mapping_df = mapping_df.dropna(subset=["key"])
        mapping_df["key"] = mapping_df["key"].str.strip().str.lower()
        # mapping_df["Categoria"] = mapping_df["Categoria"].fillna("")
        # mapping_df["Macro Categoria"] = mapping_df["Macro Categoria"].fillna("")
        mapping_df = mapping_df.drop_duplicates(subset=["key"], keep='first')

        txs_new = txs_new.rename(columns={"Categoria ": "Categoria"})
        txs_new = txs_new.rename(columns={"Categoria": "Categoria_auto"})
        txs_new['op'] = txs_new["Operazione"].str.lower().str.replace('*', '')

        txs_new["key"] = txs_new["op"].apply(
            lambda x: next((key for key in mapping_df["key"] if key in x or key == x), None)
        )

        txs_new = txs_new.merge(mapping_df, on="key", how="left")
        txs_new["cat_filling"] = "to_fill"
        txs_new.loc[txs_new["Categoria"].notna(), "cat_filling"] = "from_mapping"

        for index, row in txs_new[txs_new["cat_filling"] == "from_mapping"].iterrows():
            print("INFO:", row["Operazione"], "mapped to:", row["Categoria"], row["Macro Categoria"])

        for index, row in txs_new.iterrows():
            if row["key"]:
                continue
            operazione = row["Operazione"]

            matches = txs_old[txs_old["Operazione"].str.lower() == operazione.lower()]
            if len(matches) >= 1:
                categories = matches["Categoria"].unique().tolist()
                macro_categories = matches["Macro Categoria"].unique().tolist()
                if len(categories) == 1:
                    txs_new.at[index, "Categoria"] = categories[0]
                    txs_new.at[index, "cat_filling"] = "from_old"
                elif len(categories) > 1:
                    print("WARNING: multiple categories for", operazione)
                if len(macro_categories) == 1:
                    txs_new.at[index, "Macro Categoria"] = macro_categories[0]
                    txs_new.at[index, "cat_filling"] = "from_old"
                elif len(macro_categories) > 1:
                    print("WARNING: multiple macro categories for", operazione)
                if len(categories) == 1 and len(macro_categories) == 1:
                    # print("MATCHED ", operazione, categories[0], macro_categories[0])
                    continue

        sorted_cols = ["Data Valuta", "Operazione", "Importo", "Categoria", "Macro Categoria"]
        remaining_cols = [x for x in txs_new.columns if x not in sorted_cols]
        txs_new = txs_new[sorted_cols + remaining_cols]
        return txs_new

    def back_up_old_txs(self) -> bool:
        if self.back_up_done:
            return False
        print("Executing back-up of txs.xlsx")
        source = fr"{EXPENSES_FOLDER}\txs.xlsx"
        destination = fr"{EXPENSES_FOLDER}\backup\txs_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy(source, destination)
        self.back_up_done = True
        return True

    def update_txs(self, txs_old: pd.DataFrame, txs_new: pd.DataFrame):
        # __ load the existing workbook __
        file_path = fr"{EXPENSES_FOLDER}\txs.xlsx"
        workbook = load_workbook(file_path)

        # __ get the active sheet or specify a sheet by name __
        # sheet = workbook.active  # Use the active sheet
        sheet = workbook["txs"]  # Use a specific sheet if needed

        # __ get first empty row __
        start_row = sheet.max_row + 1

        # __ determine column positions based on the header row
        header_row = 1  # The row where headers are located (1-based index)
        header_map = {cell.value: col_idx for col_idx, cell in enumerate(sheet[header_row], start=1)}

        # __ write the DataFrame data row by row and cell by cell __
        for row_idx, (index, row) in enumerate(txs_new.iterrows(), start=start_row):
            for col_name, value in row.items():
                # Find the column index from the header map
                col_idx = header_map.get(col_name)
                if col_idx:  # Only write if the column exists in the Excel sheet
                    sheet.cell(row=row_idx, column=col_idx, value=value)

        # # Determine the last row of the original data
        # last_original_row = start_row - 1
        #
        # # Copy the formatting of the last original row to all newly added rows
        # def copy_row_format(sheet, source_row, target_row):
        #     """Copy the format of an entire row from source_row to target_row."""
        #     for col_idx, source_cell in enumerate(sheet[source_row], start=1):
        #         target_cell = sheet.cell(row=target_row, column=col_idx)
        #         if source_cell.has_style:
        #             target_cell.font = source_cell.font
        #             target_cell.border = source_cell.border
        #             target_cell.fill = source_cell.fill
        #             target_cell.number_format = source_cell.number_format
        #             target_cell.protection = source_cell.protection
        #             target_cell.alignment = source_cell.alignment
        #
        # # Apply formatting to all new rows
        # for new_row in range(start_row, sheet.max_row + 1):
        #     copy_row_format(sheet, last_original_row, new_row)

        # Save the changes to the existing file
        workbook.save(file_path)

        print("Excel file updated successfully!")

    def update_data_contabile(self, filled_txs: pd.DataFrame):
        missing = len(filled_txs)
        print("Found #", missing, "missing Data Contabile")
        filled_txs = filled_txs[filled_txs["Data Contabile"].notna()]
        file_path = fr"{EXPENSES_FOLDER}\txs.xlsx"
        workbook = load_workbook(file_path)
        sheet = workbook["txs"]
        header_row = 1  # The row where headers are located (1-based index)
        header_map = {cell.value: col_idx for col_idx, cell in enumerate(sheet[header_row], start=1)}
        column_1 = header_map["Data Valuta"] - 1
        column_2 = header_map["Operazione"] - 1
        column_3 = header_map["Importo"] - 1
        column_100 = header_map["Data Contabile"] - 1
        column_101 = header_map["Accrediti in valuta"] - 1
        column_102 = header_map["Addebiti in valuta"] - 1
        filled = 0
        for index, row in filled_txs.iterrows():
            for sheet_row in sheet.iter_rows(min_row=min(filled_txs["index"].values), max_row=sheet.max_row):
                if row["Data Valuta"] == sheet_row[column_1].value.date() and row["Operazione"] == sheet_row[column_2].value and row["Importo"] == sheet_row[column_3].value:
                    sheet_row[column_100].value = row["Data Contabile"]
                    sheet_row[column_101].value = row["Accrediti in valuta"]
                    sheet_row[column_102].value = row["Addebiti in valuta"]
                    print("FILLED", row["Data Valuta"], row["Importo"], row["Operazione"])
                    filled += 1
                    break

        print(f"Filled {filled}/{missing} Data Contabile")
        if filled > 0:
            workbook.save(file_path)
        print("Excel file updated successfully!")

    @staticmethod
    def move_files_to_old_folder():
        new_txs_files = [x for x in os.listdir(EXPENSES_FOLDER) if x.startswith('lista_completa') and x.endswith('.xlsx') and x != 'txs.xlsx']
        for f in new_txs_files:
            shutil.move(fr"{EXPENSES_FOLDER}\{f}", fr"{EXPENSES_FOLDER}\old\{f}")

        cc_details_files = [x for x in os.listdir(EXPENSES_FOLDER) if x.startswith('Movimenti_Carta') and x.endswith('.xlsx') and x != 'txs.xlsx']
        for f in cc_details_files:
            shutil.move(fr"{EXPENSES_FOLDER}\{f}", fr"{EXPENSES_FOLDER}\old\{f}")


if __name__ == '__main__':
    expenses = Expenses()
    expenses.main()
