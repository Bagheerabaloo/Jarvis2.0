import argparse
import os
import re
import json
import time
import random
from datetime import datetime
from typing import Any, Dict, Optional, Tuple

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Uses your existing Selenium wrappers (reference)
from src.common.web_driver.FirefoxDriver import FirefoxDriver
from pathlib import Path
from datetime import timedelta
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


OUTPUT_COLUMNS = [
    "index",
    "Acquisto",
    "Costo",
    "Data",
    "Categoria",
    "Negozio",
    "Marca",
    "Modello",
    "Taglia",
    "Quantità",
    "Prezzo unità",
    "Unità",
    "Link",
    "Note",
    "Order ID",
]


# -----------------------------
# Helpers: parsing & formatting
# -----------------------------
def _detect_separator(path: str) -> str:
    """Detect whether the file is tab-separated or comma-separated."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        head = f.readline()
    return "\t" if "\t" in head else ","


def _parse_amazon_datetime(date_str: str) -> Optional[datetime]:
    """Parse Amazon ISO-like datetime strings (e.g., 2026-01-02T14:16:55Z)."""
    if not date_str or not isinstance(date_str, str):
        return None
    s = date_str.strip()
    try:
        # Most Amazon exports are ISO 8601 with trailing Z
        if s.endswith("Z"):
            s = s.replace("Z", "+00:00")
        return datetime.fromisoformat(s)
    except Exception:
        # Fallback: try to extract YYYY-MM-DD
        m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
        if m:
            try:
                return datetime.fromisoformat(m.group(1))
            except Exception:
                return None
    return None


def _parse_start_date(date_str: str) -> Optional[datetime]:
    """
    Parse a start date provided by the user.
    Accepted formats:
    - YYYY-MM-DD
    - DD/MM/YYYY
    """
    if not date_str:
        return None
    s = str(date_str).strip()
    if not s:
        return None

    # YYYY-MM-DD
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return datetime(dt.year, dt.month, dt.day)
    except Exception:
        pass

    # DD/MM/YYYY
    try:
        dt = datetime.strptime(s, "%d/%m/%Y")
        return datetime(dt.year, dt.month, dt.day)
    except Exception:
        return None


def _to_float(x: Any) -> Optional[float]:
    """Convert EU/US numeric-like strings to float safely."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    # Amazon export typically uses dot decimal for EUR in CSV
    # but we also handle comma decimals just in case.
    s = s.replace("€", "").strip()
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    # Remove thousands separators (either "." or ",") conservatively
    s = re.sub(r"(?<=\d)[,\.](?=\d{3}\b)", "", s)
    try:
        return float(s)
    except Exception:
        return None


def _sleep_human(min_s: float = 0.7, max_s: float = 1.8) -> None:
    """Human-like delay to reduce bot-detection risk."""
    time.sleep(random.uniform(min_s, max_s))


def _extract_date_from_filename(filename: str) -> Optional[datetime]:
    """
    Extract a date from a filename.
    Supported patterns:
    - YYYY-MM-DD
    - YYYY_MM_DD
    - DD-MM-YYYY
    - DD_MM_YYYY
    """
    if not filename:
        return None

    s = filename

    m = re.search(r"\b(\d{8})\b", s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y%m%d")
        except Exception:
            pass

    # YYYY-MM-DD or YYYY_MM_DD
    m = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass

    # DD-MM-YYYY or DD_MM_YYYY
    m = re.search(r"(\d{2})[-_](\d{2})[-_](\d{4})", s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except Exception:
            pass

    return None


def _find_latest_backup_xlsx(backups_dir: str) -> Optional[str]:
    """
    Find the most recent .xlsx file in backups_dir.
    Primary sort: date parsed from filename (descending)
    Fallback: file modified time (descending)
    """
    if not backups_dir:
        return None

    p = Path(backups_dir)
    if not p.exists() or not p.is_dir():
        return None

    files = list(p.glob("*.xlsx"))
    if not files:
        return None

    scored = []
    for f in files:
        dt = _extract_date_from_filename(f.name)
        scored.append((dt, f.stat().st_mtime, str(f)))

    # Prefer parsed filename date if present
    with_date = [x for x in scored if x[0] is not None]
    if with_date:
        with_date.sort(key=lambda x: x[0], reverse=True)
        latest_path = with_date[0][2]
        logger.info(f"Latest backup XLSX selected: {latest_path}")
        return latest_path

    # Fallback on modified time
    scored.sort(key=lambda x: x[1], reverse=True)
    latest_path = scored[0][2]
    logger.info(f"Latest backup XLSX selected (mtime): {latest_path}")
    return latest_path


def _compute_start_date_from_latest_backup(backups_dir: str) -> Optional[datetime]:
    """
    Read latest backup XLSX and compute start date:
    max(Data) - 1 day
    """
    latest = _find_latest_backup_xlsx(backups_dir)
    if not latest:
        return None

    try:
        prev = pd.read_excel(latest, sheet_name="Acquisti Passati", engine="openpyxl")
        if "Data" not in prev.columns:
            return None
        if "Negozio" not in prev.columns:
            return None

        prev = prev[prev["Negozio"].astype(str).str.strip().str.lower() == "amazon"]
        if prev.empty:
            return None

        # Accept Excel dates or strings like DD/MM/YYYY
        d = pd.to_datetime(prev["Data"], errors="coerce", dayfirst=True)
        d = d.dropna()
        if d.empty:
            return None

        last_dt = d.max().to_pydatetime()
        last_day = datetime(last_dt.year, last_dt.month, last_dt.day)
        computed_dt = last_day - timedelta(days=1)
        logger.info(f"Computed start date from backup (max(Data)-1): {computed_dt.strftime('%Y-%m-%d')}")
        return computed_dt
    except Exception:
        return None


def _find_latest_orders_folder(root_dir: str) -> Path:
    """
    root_dir/
      amazon/
        2026-01-10 - Your Orders/
        2026-01-02 - Your Orders/
    Picks the folder with the most recent YYYY-MM-DD prefix.
    """
    base = Path(root_dir)
    amazon_dir = base / "Amazon"
    if not amazon_dir.exists():
        # fallback: try "Amazon"
        amazon_dir = base / "Amazon"

    if not amazon_dir.exists() or not amazon_dir.is_dir():
        raise FileNotFoundError(f"amazon folder not found under: {base}")

    candidates = []
    for p in amazon_dir.iterdir():
        if not p.is_dir():
            continue
        m = re.match(r"^(\d{4}-\d{2}-\d{2})\s+-\s+Your Orders$", p.name)
        if not m:
            continue
        try:
            dt = datetime.strptime(m.group(1), "%Y-%m-%d")
            candidates.append((dt, p))
        except Exception:
            continue

    if not candidates:
        raise FileNotFoundError(f"No 'YYYY-MM-DD - Your Orders' folders found in: {amazon_dir}")

    candidates.sort(key=lambda x: x[0], reverse=True)
    logger.info(f"Latest 'Your Orders' folder selected: {candidates[0][1]}")
    return candidates[0][1]


def _resolve_amazon_csv_from_root(root_dir: str) -> tuple[str, str]:
    """
    Returns (input_csv_path, output_xlsx_path) based on latest orders folder.
    Expects:
      <latest>/Retail.OrderHistory.2/Retail.OrderHistory.2.csv
    Output:
      <latest>/output.xlsx
    """
    latest_dir = _find_latest_orders_folder(root_dir)

    csv_path = latest_dir / "Retail.OrderHistory.2" / "Retail.OrderHistory.2.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"Retail.OrderHistory.2.csv not found at: {csv_path}")

    output_path = latest_dir / "output.xlsx"
    logger.info(f"Amazon CSV resolved: {csv_path}")
    logger.info(f"Output XLSX resolved: {output_path}")
    return str(csv_path), str(output_path)


def _book_title_from_product_name(name: str) -> str:
    """Extract a clean book title from Amazon product name (best-effort)."""
    s = (name or "").strip()
    s = re.sub(r"\(.*?\)", "", s).strip()  # remove parenthetical
    # remove common format words
    s = re.sub(r"\b(kindlekindle|kindle|copertina\s+\w+|edizione|paperback|hardcover)\b.*", "", s, flags=re.IGNORECASE).strip()
    # take first segment (often the title)
    parts = re.split(r"\s*[,\-|–]\s*", s)
    title = (parts[0] if parts else s).strip()
    return title[:100]


def _is_amazon_fresh(carrier_tracking: str) -> bool:
    """Detect Amazon Fresh deliveries based on carrier/tracking pattern."""
    s = (carrier_tracking or "").upper()
    return "RABBIT(" in s


def _parse_count_from_size(text: str) -> Optional[Tuple[float, str]]:
    """
    Parse count-like sizes such as:
    - "100 pezzi", "48 pastiglie", "24 compresse", "12 rotoli", "132 lavaggi"
    Returns (count, unit_label) if found.
    """
    s = (text or "").strip().lower()
    if not s:
        return None

    m = re.search(r"(\d+(?:[.,]\d+)?)\s*(pezzi|pz\.?|pastiglie|compresse|rotoli|capsule|lavaggi|fogli|buste|sacchi|filtri)\b", s)
    if not m:
        return None

    raw_num = m.group(1).replace(",", ".")
    try:
        count = float(raw_num)
    except Exception:
        return None

    unit = m.group(2)
    # Normalize unit labels
    if unit.startswith("pz") or unit == "pezzi":
        unit = "pz"
    return count, unit


def _parse_liters_from_size(text: str) -> Optional[float]:
    """
    Parse volume sizes and convert to liters:
    - "150ml" -> 0.150
    - "1L" or "1,2L" -> 1.2
    - "1760 ml x 3" -> 1.760 * 3
    Returns liters if found, else None.
    """
    s = (text or "").strip().lower()
    if not s:
        return None

    # With multiplier: "1760 ml x 3"
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*(ml|l)\s*(?:x|×)\s*(\d+)", s)
    if m:
        val = float(m.group(1).replace(",", "."))
        unit = m.group(2)
        mult = int(m.group(3))
        liters = (val / 1000.0) if unit == "ml" else val
        return liters * mult

    # Single volume: "500ml" or "1,5l"
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*(ml|l)\b", s)
    if not m:
        return None

    val = float(m.group(1).replace(",", "."))
    unit = m.group(2)
    liters = (val / 1000.0) if unit == "ml" else val
    return liters


def _normalize_qty(x: float) -> Any:
    """Return int if it's an integer, else a rounded float (keeps Excel numeric)."""
    if x is None:
        return ""
    if abs(x - round(x)) < 1e-9:
        return int(round(x))
    return round(float(x), 3)


def _derive_qty_unit_and_unit_price(
    base_qty: float,
    total_owed: Optional[float],
    size_text: str,
) -> Tuple[Any, str, Optional[float]]:
    """
    Decide how to interpret quantity based on size_text:
    - If size is count-based (e.g. '100 pezzi'): qty = base_qty * 100, unit = 'pz' (or pastiglie/compresse/...)
    - If size is volume-based (e.g. '150ml'): qty = base_qty * 0.150, unit = 'L', unit_price = €/L
    - Otherwise: qty = base_qty, unit = 'pz' (piece-based), unit_price = €/pz
    """
    qty_out = float(base_qty) if base_qty else 1.0
    unit_out = "pz"

    # 1) Count-based
    parsed_count = _parse_count_from_size(size_text)
    if parsed_count:
        per_pack, unit_label = parsed_count
        qty_out = qty_out * per_pack
        unit_out = unit_label

    else:
        # 2) Volume-based (€/L)
        liters = _parse_liters_from_size(size_text)
        if liters is not None:
            qty_out = qty_out * liters
            unit_out = "L"

    unit_price = None
    if total_owed is not None and qty_out:
        unit_price = total_owed / qty_out

    return _normalize_qty(qty_out), unit_out, unit_price



# -----------------------------
# LLM Client (DeepSeek OpenAI-compatible)
# -----------------------------
class DeepSeekLLM:
    def __init__(self):
        self.api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
        self.base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1").strip()
        self.model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat").strip()

        # --- NEW counters ---
        self.calls_single = 0
        self.calls_batch = 0

        self.log_json = os.getenv("DEEPSEEK_LOG_JSON", "1").strip() == "1"
        self.log_raw = os.getenv("DEEPSEEK_LOG_RAW", "0").strip() == "1"
        self.log_max_chars = int(os.getenv("DEEPSEEK_LOG_MAX_CHARS", "3000"))

    def enabled(self) -> bool:
        return bool(self.api_key)

    def _log_deepseek_output(self, label: str, raw: str, parsed: Any) -> None:
        if self.log_raw:
            logger.info(f"[DeepSeek][{label}] RAW (truncated): {raw[:self.log_max_chars]}")
        if self.log_json:
            try:
                logger.info(f"[DeepSeek][{label}] PARSED_JSON (truncated): {json.dumps(parsed, ensure_ascii=False)[:self.log_max_chars]}")
            except Exception:
                logger.info(f"[DeepSeek][{label}] PARSED_JSON: <non-serializable>")

    def enrich_product(
            self,
            product_name: str,
            scraped_brand: str = "",
            scraped_model: str = "",
            scraped_size: str = "",
            amazon_title: str = "",
            amazon_variations: Optional[dict] = None,
    ) -> Dict[str, str]:
        """
        Return a JSON-like dict with:
        - acquisto: short purchase name (Italian, 2-6 words)
        - categoria: a short category label (Italian, 1-2 words)
        - modello: model/version if present (else empty)
        - note: optional (else empty)

        If LLM is not enabled or fails, returns heuristic output.
        """
        if not self.enabled():
            return self._fallback(product_name, scraped_brand, scraped_model, scraped_size)

        system_msg = (
            "You output ONLY valid JSON. No prose. "
            "Keys: acquisto, categoria, modello, taglia, note. "
            "Values must be strings. If unknown, use empty string."
        )

        variations = amazon_variations or {}
        variations_json = json.dumps(variations, ensure_ascii=False)

        user_msg = f"""
        INPUT FIELDS (may be empty):
        - product_name: {product_name}
        - scraped_brand: {scraped_brand}
        - scraped_model: {scraped_model}
        - scraped_size: {scraped_size}
        - amazon_title: {amazon_title}
        - variations_json: {variations_json}

        Return ONLY JSON with keys: acquisto, categoria, modello, taglia, note.

        GOAL:
        Create a short Italian purchase label optimized for a shopping sheet.

        acquisto:
        1) Format: "<Tipo> <Marca> <Modello/Linea>" with 2–5 words.
           - If you do not have a model/line, use "<Tipo> <Marca>".
        2) 'Marca' priority:
           - Use scraped_brand if provided.
           - Otherwise infer brand from product_name if clearly present.
        3) 'Modello/Linea' priority:
           - Use scraped_model if it contains something meaningful.
           - Otherwise extract the shortest distinctive model/line from product_name (e.g., "Revolution 7").
        4) Avoid generic/marketing fluff: remove words like "set", "cofanetto", "idee regalo", "profumo sensuale",
           "confezione", "pezzi", "ml", "latta", "anniversario", etc.
        5) If the product is LEGO:
           - 'acquisto' must start with "LEGO" and include the set name if available (e.g., "LEGO Van Gogh").
        6) If categoria is "Libri":
           - acquisto must be exactly "Libro - <titolo pulito>" (remove format words like copertina, kindle, edizione).
        7) Keep 'acquisto' under 35 characters when possible.
        
        modello (IMPORTANT):
        - Must be the product line/version/name, not a quantity/format.
        - Priority: scraped_model > amazon_title > product_name.
        - Good models: "Revolution 7 Uomo", "MX Keys S", "Mini 4K", "FRAKTA", "Powerball Ultimate PLUS".
        - BAD as modello (these belong to taglia): "1L", "500ml", "132 lavaggi", "48 pastiglie", "2m", "175 grammi", "128GB".
        - If both a real model and a format exist, keep model in modello and put the format in taglia.

        taglia (IMPORTANT):
        Extract ONE primary size/format, normalized, using (variations_json > scraped_size > amazon_title > product_name).
        Rules by type:
        - Clothing/shoes: sizes like XS/S/M/L/XL or EU 42 or ranges "40-42".
        - Cables: length like "1m", "2m".
        - Liquids: "500ml", "1L", "1760ml".
        - Capacity/storage: "128GB", "2TB".
        - Power: "27W", "150W".
        - Consumables: "48 pastiglie", "24 compresse", "12 rotoli", "132 lavaggi", "10 pezzi", "175 grammi".
        Pick the most informative single one. Do NOT return multiple formats unless they are fused like "1760ml x3".
        
        categoria:
        Choose one of: Libri, PC, Casa, Bellezza, Elettronica, Abbigliamento, Alimentari, Ufficio, Animali, Altro.
        
        Examples:
        1) Nike shoes
        product_name: "Nike Fb2207 Revolution 7 Uomo ... EU 42"
        -> modello: "Revolution 7 Uomo", taglia: "42", acquisto: "Scarpe Nike Revolution 7"
        
        2) Detergent
        amazon_title: "Omino Bianco ... 132 Lavaggi ... 1760 ml x 3"
        -> modello: "Igienizzante 132 lavaggi x3", taglia: "1760ml", acquisto: "Detersivo Omino Bianco"
        
        3) Cable
        product_name: "UGREEN Cavo HDMI 2.1 ... (2M)"
        -> modello: "HDMI 2.1", taglia: "2m", acquisto: "Cavo HDMI UGREEN"
        
        Return ONLY JSON.
        """.strip()

        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
            "temperature": 0.2,
        }

        url = f"{self.base_url.rstrip('/')}/chat/completions"
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }

        self.calls_single += 1
        call_n = self.calls_single
        t0 = time.perf_counter()
        payload_bytes = len(json.dumps(payload, ensure_ascii=False).encode("utf-8"))
        logger.info(
            f"[DEEPSEEK] single#{call_n} POST {url} model={self.model} payload_bytes={payload_bytes} "
            f"product='{(product_name or '')[:80]}'"
        )

        try:
            r = requests.post(url, headers=headers, json=payload, timeout=30)
            r.raise_for_status()

            elapsed = time.perf_counter() - t0
            logger.info(
                f"[DEEPSEEK] single#{call_n} status={r.status_code} time={elapsed:.2f}s resp_bytes={len(r.content)}"
            )

            data = r.json()
            content = data["choices"][0]["message"]["content"].strip()
            obj = self._safe_json_loads(content)

            self._log_deepseek_output(
                label="single",
                raw=content,
                parsed=obj
            )

            if obj and isinstance(obj, dict):
                return {
                    "acquisto": str(obj.get("acquisto", "") or ""),
                    "categoria": str(obj.get("categoria", "") or ""),
                    "modello": str(obj.get("modello", "") or ""),
                    "taglia": str(obj.get("taglia", "") or ""),
                    "note": str(obj.get("note", "") or ""),
                }
        except Exception as e:
            logger.warning(f"[DEEPSEEK] single#{call_n} ERROR: {e}")

        return self._fallback(product_name)

    def enrich_products_batch(self, items: list[dict]) -> dict:
        """
        items: [{ "id": "...", "product_name": "...", "brand": "...", "model": "...", "size": "...", "title": "...", "variations": {...} }]
        returns: { id: {acquisto,categoria,modello,taglia,note} }
        """
        if not self.enabled():
            return {}

        system_msg = (
            "You output ONLY valid JSON. No prose. "
            "Return a JSON array of objects. "
            "Each object must contain: id, acquisto, categoria, modello, taglia, note. "
            "All values must be strings."
        )

        # Keep variations compact
        compact_items = []
        for it in items:
            compact_items.append({
                "id": it.get("id", ""),
                "product_name": it.get("product_name", ""),
                "scraped_brand": it.get("scraped_brand", ""),
                "scraped_model": it.get("scraped_model", ""),
                "scraped_size": it.get("scraped_size", ""),
                "amazon_title": it.get("amazon_title", ""),
                "variations_json": it.get("variations_json", {}),
            })

        user_msg = f"""
        YOU WILL RECEIVE A LIST OF PRODUCTS AS JSON.
        For EACH item you MUST output one JSON object in a JSON ARRAY.
        
        Each input item has these fields (may be empty):
        - id
        - product_name
        - scraped_brand
        - scraped_model
        - scraped_size
        - amazon_title
        - variations_json
        
        Return ONLY a JSON ARRAY (no prose, no markdown).
        Each output object MUST contain:
        id, acquisto, categoria, modello, taglia, note
        All values must be strings. If unknown, use empty string.
        The "id" must be copied exactly from input.

        GOAL:
        Create a short Italian purchase label optimized for a shopping sheet.

        acquisto:
        1) Format: "<Tipo> <Marca> <Modello/Linea>" with 2–5 words.
           - If you do not have a model/line, use "<Tipo> <Marca>".
        2) 'Marca' priority:
           - Use scraped_brand if provided.
           - Otherwise infer brand from product_name if clearly present.
        3) 'Modello/Linea' priority:
           - Use scraped_model if it contains something meaningful.
           - Otherwise extract the shortest distinctive model/line from product_name (e.g., "Revolution 7").
        4) Avoid generic/marketing fluff: remove words like "set", "cofanetto", "idee regalo", "profumo sensuale",
           "confezione", "pezzi", "ml", "latta", "anniversario", etc.
        5) If the product is LEGO:
           - 'acquisto' must start with "LEGO" and include the set name if available (e.g., "LEGO Van Gogh").
        6) If categoria is "Libri":
           - acquisto must be exactly "Libro - <titolo pulito>" (remove format words like copertina, kindle, edizione).
        7) Keep 'acquisto' under 35 characters when possible.

        modello (IMPORTANT):
        - Must be the product line/version/name, not a quantity/format.
        - Priority: scraped_model > amazon_title > product_name.
        - Good models: "Revolution 7 Uomo", "MX Keys S", "Mini 4K", "FRAKTA", "Powerball Ultimate PLUS".
        - BAD as modello (these belong to taglia): "1L", "500ml", "132 lavaggi", "48 pastiglie", "2m", "175 grammi", "128GB".
        - If both a real model and a format exist, keep model in modello and put the format in taglia.

        taglia (IMPORTANT):
        Extract ONE primary size/format, normalized, using (variations_json > scraped_size > amazon_title > product_name).
        Rules by type:
        - Clothing/shoes: sizes like XS/S/M/L/XL or EU 42 or ranges "40-42".
        - Cables: length like "1m", "2m".
        - Liquids: "500ml", "1L", "1760ml".
        - Capacity/storage: "128GB", "2TB".
        - Power: "27W", "150W".
        - Consumables: "48 pastiglie", "24 compresse", "12 rotoli", "132 lavaggi", "10 pezzi", "175 grammi".
        Pick the most informative single one. Do NOT return multiple formats unless they are fused like "1760ml x3".

        categoria:
        Choose one of: Libri, PC, Casa, Bellezza, Elettronica, Abbigliamento, Alimentari, Ufficio, Animali, Altro.

        Examples:
        1) Nike shoes
        product_name: "Nike Fb2207 Revolution 7 Uomo ... EU 42"
        -> modello: "Revolution 7 Uomo", taglia: "42", acquisto: "Scarpe Nike Revolution 7"

        2) Detergent
        amazon_title: "Omino Bianco ... 132 Lavaggi ... 1760 ml x 3"
        -> modello: "Igienizzante 132 lavaggi x3", taglia: "1760ml", acquisto: "Detersivo Omino Bianco"

        3) Cable
        product_name: "UGREEN Cavo HDMI 2.1 ... (2M)"
        -> modello: "HDMI 2.1", taglia: "2m", acquisto: "Cavo HDMI UGREEN"

        NOW PROCESS THIS LIST AND RETURN ONLY THE JSON ARRAY:
        
        ITEMS_JSON:
        {json.dumps(compact_items, ensure_ascii=False)}
        """.strip()

        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
            "temperature": 0.2,
        }

        url = f"{self.base_url.rstrip('/')}/chat/completions"
        headers = {"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"}

        self.calls_batch += 1
        call_n = self.calls_batch
        t0 = time.perf_counter()
        payload_bytes = len(json.dumps(payload, ensure_ascii=False).encode("utf-8"))
        logger.info(
            f"[DEEPSEEK] batch#{call_n} items={len(items)} POST {url} model={self.model} payload_bytes={payload_bytes}"
        )

        try:
            r = requests.post(url, headers=headers, json=payload, timeout=60)
            r.raise_for_status()
        except Exception as e:
            logger.warning(f"[DEEPSEEK] batch#{call_n} HTTP ERROR: {e}")
            return {}

        elapsed = time.perf_counter() - t0
        logger.info(
            f"[DEEPSEEK] batch#{call_n} status={r.status_code} time={elapsed:.2f}s resp_bytes={len(r.content)}"
        )

        data = r.json()
        content = data["choices"][0]["message"]["content"].strip()
        arr = self._safe_json_loads(content)

        self._log_deepseek_output(
            label=f"batch_{len(items)}",
            raw=content,
            parsed=arr
        )

        # Expect list[dict]
        out = {}
        if isinstance(arr, list):
            for obj in arr:
                if isinstance(obj, dict) and obj.get("id"):
                    out[str(obj["id"])] = {
                        "acquisto": str(obj.get("acquisto", "") or ""),
                        "categoria": str(obj.get("categoria", "") or ""),
                        "modello": str(obj.get("modello", "") or ""),
                        "taglia": str(obj.get("taglia", "") or ""),
                        "note": str(obj.get("note", "") or ""),
                    }

        logger.info(f"[DEEPSEEK] batch#{call_n} parsed_ok={len(out)}/{len(items)}")

        return out

    @staticmethod
    def _safe_json_loads(text: str) -> Optional[Any]:
        """Try to parse JSON (object or array), also if wrapped in fences or extra text."""
        if not text:
            return None

        # Remove ```json fences if present
        t = text.strip()
        t = re.sub(r"^```(?:json)?\s*", "", t)
        t = re.sub(r"\s*```$", "", t)

        # Try direct parse
        try:
            return json.loads(t)
        except Exception:
            pass

        # Try to extract a JSON array first
        m = re.search(r"\[.*\]", t, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                pass

        # Fallback: try to extract first JSON object
        m = re.search(r"\{.*\}", t, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                return None

        return None

    @staticmethod
    def _fallback(product_name: str, scraped_brand: str = "", scraped_model: str = "", scraped_size: str = "") -> Dict[str, str]:
        """Simple heuristic if no LLM available."""
        s = (product_name or "").strip()
        short = s.split(",")[0].strip() if s else ""
        # Minimal category guess
        low = s.lower()
        if any(k in low for k in ["libro", "kindle", "romanzo", "edizione"]):
            cat = "Libri"
        elif any(k in low for k in ["shampoo", "profumo", "eau de toilette", "crema", "cosmet", "bagnodoccia"]):
            cat = "Bellezza"
        elif any(k in low for k in ["pc", "ssd", "ram", "mouse", "tastiera", "monitor"]):
            cat = "PC"
        else:
            cat = "Altro"

        return {"acquisto": short[:60], "categoria": cat, "modello": "", "taglia": "", "note": ""}


# -----------------------------
# Amazon Scraper (ASIN-first)
# -----------------------------
class AmazonScraper:
    def __init__(self, headless: bool = True, is_raspberry: bool = False):
        # Same settings as AutoScout:
        # os_environ=is_raspberry, headless=headless, selenium_profile=True
        self.driver = FirefoxDriver(
            os_environ=is_raspberry,
            headless=headless,
            selenium_profile=True,
        )
        self.cache = {}

        # --- NEW: counters ---
        self.page_visits = 0       # visite REALI (non cache)
        self.cache_hits = 0

    def start(self) -> None:
        """Start Selenium driver."""
        self.driver.init_driver()

    def close(self) -> None:
        """Close Selenium driver."""
        self.driver.close_driver()

    def get_product_info(self, asin: str, fallback_query: str = "") -> Dict[str, str]:
        """
        Scrape Amazon product page via ASIN: https://www.amazon.it/dp/{asin}
        Returns dict: brand, size, model, link.
        """
        asin = (asin or "").strip()

        # --- avoid caching empty asin ---
        if asin and asin in self.cache:
            self.cache_hits += 1
            logger.debug(f"[SCRAPER] cache hit asin={asin} (hits={self.cache_hits})")
            return self.cache[asin]

        info = {"title": "", "brand": "", "size": "", "model": "", "link": ""}

        # If no ASIN, don't visit
        if not asin:
            return info

        url = f"https://www.amazon.it/dp/{asin}"
        info["link"] = url

        # --- NEW: page counter + log start ---
        self.page_visits += 1
        page_n = self.page_visits
        t0 = time.perf_counter()
        logger.info(f"[SCRAPER] page={page_n} GET {url} asin={asin}")

        try:
            self.driver.get_url(url, do_wait=True, add_slash=False)
            _sleep_human()
            self._best_effort_accept_cookies()
            soup = self.driver.get_response()
            info["variations"] = self._extract_variations(soup)

            title = self._text(soup.select_one("#productTitle"))
            info["title"] = title
            byline = self._text(soup.select_one("#bylineInfo"))
            tables = self._extract_details_tables(soup)
            bullets = self._extract_detail_bullets(soup)

            # Brand: prefer byline store
            info["brand"] = self._extract_brand_from_byline(byline) or self._pick_value(
                tables, bullets, keys=["Marca", "Brand", "Produttore"]
            )

            # Model: from product details if present
            info["model"] = self._pick_value(
                tables, bullets, keys=["Nome modello", "Modello", "Numero modello articolo", "Model"]
            )

            # Size: try variations first, then details/bullets
            info["size"] = self._extract_variation_size(soup) or self._pick_value(
                tables, bullets, keys=["Taglia", "Misura", "Dimensioni", "Formato"]
            )

            # Link: if we ended up on a canonical URL, use current_url
            try:
                info["link"] = self.driver.driver.current_url or info["link"]
            except Exception:
                pass

            # --- NEW: log end ---
            elapsed = time.perf_counter() - t0
            blocked = "YES" if not title else "NO"
            logger.info(
                f"[SCRAPER] page={page_n} DONE {elapsed:.2f}s blocked={blocked} "
                f"title_len={len(title)} brand='{info['brand']}' size='{info['size']}'"
            )

        except Exception as e:
            elapsed = time.perf_counter() - t0
            logger.warning(f"[SCRAPER] page={page_n} ERROR after {elapsed:.2f}s asin={asin}: {e}")

        self.cache[asin] = info
        return info

    def _best_effort_accept_cookies(self) -> None:
        """Try common Amazon cookie consent buttons (best-effort)."""
        try:
            # Common IDs in EU consent banners
            btn = None
            for css in ["#sp-cc-accept", "input#sp-cc-accept", "button#sp-cc-accept"]:
                try:
                    btn = self.driver.driver.find_element("css selector", css)
                    if btn:
                        btn.click()
                        _sleep_human(0.3, 0.8)
                        return
                except Exception:
                    continue
        except Exception:
            return

    @staticmethod
    def _text(el) -> str:
        return el.get_text(" ", strip=True) if el else ""

    @staticmethod
    def _extract_brand_from_byline(byline: str) -> str:
        """
        Examples:
        - 'Visita lo Store di XYZ'
        - 'Marca: XYZ'
        """
        if not byline:
            return ""
        m = re.search(r"Store di\s+(.+)$", byline.strip(), re.IGNORECASE)
        if m:
            return m.group(1).strip()
        m2 = re.search(r"Marca:\s*(.+)$", byline.strip(), re.IGNORECASE)
        if m2:
            return m2.group(1).strip()
        return ""

    @staticmethod
    def _extract_variation_size(soup) -> str:
        """
        Try common Amazon variation blocks for size.
        """
        # Example: <div id="variation_size_name"> ... <span class="selection">XL</span>
        sel = soup.select_one("#variation_size_name .selection")
        if sel and sel.get_text(strip=True):
            return sel.get_text(" ", strip=True)

        # Some pages use dropdown prompt
        sel = soup.select_one("#variation_size_name .a-dropdown-prompt")
        if sel and sel.get_text(strip=True):
            return sel.get_text(" ", strip=True)

        # Another common dimension label
        sel = soup.select_one("#inline-twister-expanded-dimension-text")
        if sel and sel.get_text(strip=True):
            return sel.get_text(" ", strip=True)

        return ""

    @staticmethod
    def _extract_details_tables(soup) -> Dict[str, str]:
        """
        Extract key/value pairs from Amazon product details tables.
        """
        out: Dict[str, str] = {}

        # Tech specs table
        for tr in soup.select("table#productDetails_techSpec_section_1 tr"):
            th = tr.select_one("th")
            td = tr.select_one("td")
            k = th.get_text(" ", strip=True) if th else ""
            v = td.get_text(" ", strip=True) if td else ""
            if k and v:
                out[k] = v

        # Detail bullets table
        for tr in soup.select("table#productDetails_detailBullets_sections1 tr"):
            th = tr.select_one("th")
            td = tr.select_one("td")
            k = th.get_text(" ", strip=True) if th else ""
            v = td.get_text(" ", strip=True) if td else ""
            if k and v:
                out[k] = v

        return out

    @staticmethod
    def _extract_detail_bullets(soup) -> Dict[str, str]:
        """
        Extract key/value from 'detailBullets' list (best-effort).
        """
        out: Dict[str, str] = {}
        for li in soup.select("#detailBullets_feature_div li"):
            txt = li.get_text(" ", strip=True)
            # Often formatted like: "Marca : XYZ"
            if ":" in txt:
                parts = [p.strip() for p in txt.split(":", 1)]
                if len(parts) == 2 and parts[0] and parts[1]:
                    out[parts[0]] = parts[1]
        return out

    @staticmethod
    def _pick_value(
        tables: Dict[str, str],
        bullets: Dict[str, str],
        keys: list[str],
    ) -> str:
        """
        Pick the first matching value among normalized keys from tables or bullets.
        """
        def norm(s: str) -> str:
            return re.sub(r"\s+", " ", (s or "")).strip().lower()

        t_norm = {norm(k): v for k, v in (tables or {}).items()}
        b_norm = {norm(k): v for k, v in (bullets or {}).items()}

        for k in keys:
            nk = norm(k)
            if nk in t_norm and t_norm[nk]:
                return t_norm[nk]
            if nk in b_norm and b_norm[nk]:
                return b_norm[nk]
        return ""

    @staticmethod
    def _extract_variations(soup) -> dict:
        """Extract selected Amazon variation values (size/color/style/capacity/etc.)."""
        out = {}
        for div in soup.select("div[id^='variation_'][id$='_name']"):
            var_id = (div.get("id") or "").strip()  # e.g., variation_size_name
            key = var_id.replace("variation_", "").replace("_name", "")
            sel = div.select_one(".selection") or div.select_one(".a-dropdown-prompt")
            val = sel.get_text(" ", strip=True) if sel else ""
            if key and val:
                out[key] = val
        return out

# -----------------------------
# XLSX writing
# -----------------------------
def write_xlsx(rows: list[dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acquisti"

    # Header
    ws.append(OUTPUT_COLUMNS)
    for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)

    # Rows
    for r in rows:
        ws.append([r.get(col, "") for col in OUTPUT_COLUMNS])

    # Column widths (simple, readable)
    widths = {
        "A": 7,   # index
        "B": 32,  # Acquisto
        "C": 12,  # Costo
        "D": 12,  # Data
        "E": 14,  # Categoria
        "F": 12,  # Negozio
        "G": 18,  # Marca
        "H": 18,  # Modello
        "I": 14,  # Taglia
        "J": 10,  # Quantità
        "K": 14,  # Prezzo unità
        "L": 10,  # Unità
        "M": 45,  # Link
        "N": 28,  # Note
        "O": 18,  # Order ID
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Formats: Costo, Prezzo unità as EUR; Data as dd/mm/yyyy
    # Find column indices
    col_map = {name: idx + 1 for idx, name in enumerate(OUTPUT_COLUMNS)}
    cost_col = col_map["Costo"]
    unit_price_col = col_map["Prezzo unità"]
    date_col = col_map["Data"]

    for row_idx in range(2, ws.max_row + 1):
        # Currency formatting
        for c in [cost_col, unit_price_col]:
            cell = ws.cell(row=row_idx, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#.##0,00'

        # Date formatting
        dcell = ws.cell(row=row_idx, column=date_col)
        if isinstance(dcell.value, datetime):
            dcell.number_format = "DD/MM/YYYY"

    wb.save(out_path)
    logger.info(f"XLSX saved: {out_path}")


# -----------------------------
# Main transform
# -----------------------------
def transform(
    input_csv: str,
    output_xlsx: str,
    headless: bool,
    use_llm: bool,
    use_scraping: bool,
    max_rows: Optional[int],
    is_raspberry: bool = False,
    start_date: str = "",
    backups_dir: str = "",
) -> None:

    sep = _detect_separator(input_csv)
    df = pd.read_csv(input_csv, sep=sep, dtype=str, keep_default_na=False)

    logger.info(f"Rows loaded from CSV: {len(df)}")
    logger.info(f"LLM enabled: {use_llm}")
    logger.info(f"Scraping enabled: {use_scraping}")

    # start_date has priority (override)
    start_dt = _parse_start_date(start_date)

    # If not provided, compute it from latest backup XLSX
    if not start_dt and backups_dir:
        start_dt = _compute_start_date_from_latest_backup(backups_dir)

    if _parse_start_date(start_date):
        logger.info(f"Start date override from CLI: {start_dt.strftime('%Y-%m-%d')}")
    elif start_dt:
        logger.info(f"Start date from backups: {start_dt.strftime('%Y-%m-%d')}")
    else:
        logger.info("Start date: [none] (no filtering by date)")

    if start_dt:
        before = len(df)

        # Parse Order Date column once, then filter rows before processing
        df["_order_dt"] = pd.to_datetime(df["Order Date"], errors="coerce", utc=True)

        # Keep only rows with a valid date AND date >= start_dt (day granularity)
        start_ts = pd.Timestamp(start_dt.date(), tz="UTC")
        df = df[df["_order_dt"].notna()]
        df = df[df["_order_dt"].dt.normalize() >= start_ts]

        after = len(df)
        logger.info(f"Rows after date filter: {after} (removed {before - after})")
    else:
        # Still parse once so the loop can rely on it
        df["_order_dt"] = pd.to_datetime(df["Order Date"], errors="coerce", utc=True)

    # -----------------------------
    # Init LLM + Scraper
    # -----------------------------
    llm = DeepSeekLLM()
    if use_llm and not llm.enabled():
        logger.warning("LLM requested but DEEPSEEK_API_KEY is missing/empty -> disabling LLM.")
        use_llm = False

    scraper = AmazonScraper(headless=headless, is_raspberry=is_raspberry) if use_scraping else None
    if scraper:
        scraper.start()
        logger.info("Scraper started.")

    # -----------------------------
    # Batch buffers (Amazon vs Amazon Fresh)
    # -----------------------------
    batch_size = 10  # You can tune this (10-20 recommended)
    pending_amazon: list[dict] = []
    pending_fresh: list[dict] = []
    row_by_id: dict[str, dict] = {}

    out_rows: list[dict] = []

    def _order_day_from_row(row_: pd.Series) -> Any:
        """Return datetime(date) or empty string."""
        ts = row_.get("_order_dt", None)
        try:
            if isinstance(ts, pd.Timestamp) and not pd.isna(ts):
                dt_py = ts.to_pydatetime()
                return datetime(dt_py.year, dt_py.month, dt_py.day)
        except Exception:
            pass

        dt_fallback = _parse_amazon_datetime(row_.get("Order Date", "") or "")
        if dt_fallback:
            return datetime(dt_fallback.year, dt_fallback.month, dt_fallback.day)
        return ""

    def _apply_enrich(item_id: str, enrich: dict) -> None:
        """Apply LLM enrichment to the already-created output row and recompute qty/unit/unit_price."""
        r = row_by_id.get(item_id)
        if not r:
            return

        product_name = r.get("_product_name", "") or ""
        model_from_page = r.get("_model_from_page", "") or ""
        scraped_size = r.get("_scraped_size", "") or ""
        amazon_title = r.get("_amazon_title", "") or ""
        base_qty = float(r.get("_base_qty", 1) or 1)
        total_owed = r.get("_total_owed", None)

        # Extract fields from enrich (strings)
        acquisto = (enrich or {}).get("acquisto", "") or ""
        categoria = (enrich or {}).get("categoria", "") or ""
        modello = (enrich or {}).get("modello", "") or ""
        taglia = (enrich or {}).get("taglia", "") or ""

        # Fallbacks
        if not acquisto:
            acquisto = (product_name.split(",")[0].strip() if product_name else "")
        if not categoria:
            categoria = "Altro"

        # Books rule (force)
        if categoria.strip().lower() == "libri":
            title = _book_title_from_product_name(product_name)
            acquisto = f"Libro - {title}" if title else "Libro"

        # Model fallback to scraped model
        if not modello and model_from_page:
            modello = model_from_page

        # Size fallback to scraped size
        if not taglia and scraped_size:
            taglia = scraped_size

        # Update row fields
        r["Acquisto"] = acquisto
        r["Categoria"] = categoria
        r["Modello"] = modello
        r["Taglia"] = taglia

        # Recompute quantity/unit/unit_price based on best size text
        size_text_for_calc = taglia or scraped_size or amazon_title or product_name
        qty_out, unita_out, unit_price = _derive_qty_unit_and_unit_price(
            base_qty=base_qty,
            total_owed=total_owed,
            size_text=size_text_for_calc,
        )
        r["Quantità"] = qty_out
        r["Unità"] = unita_out
        r["Prezzo unità"] = float(unit_price) if unit_price is not None else ""

    def _flush_batch(pending: list[dict], label: str) -> None:
        """Send a batch to LLM and apply results."""
        if not use_llm or not pending:
            return
        logger.info(f"LLM batch call [{label}] size={len(pending)}")
        try:
            enriched_map = llm.enrich_products_batch(pending)  # {id: {...}}
        except Exception as e:
            logger.warning(f"LLM batch failed [{label}]: {e}")
            return

        if not isinstance(enriched_map, dict) or not enriched_map:
            logger.warning(f"LLM batch returned empty/invalid map [{label}]")
            return

        for _id, _enrich in enriched_map.items():
            _apply_enrich(str(_id), _enrich if isinstance(_enrich, dict) else {})

    try:
        processed = 0

        for i, row in df.iterrows():
            if max_rows is not None and processed >= max_rows:
                break
            processed += 1

            product_name = row.get("Product Name", "") or ""
            asin = row.get("ASIN", "") or ""
            order_id = row.get("Order ID", "") or ""
            carrier_tracking = row.get("Carrier Name & Tracking Number", "") or ""
            negozio = "Amazon Fresh" if _is_amazon_fresh(carrier_tracking) else "Amazon"
            base_qty = _to_float(row.get("Quantity", "1")) or 1
            total_owed = _to_float(row.get("Total Owed", ""))  # Prices must come ONLY from CSV Total Owed
            order_day = _order_day_from_row(row)

            # Scraping enrichment (brand, size, link, plus a model hint)
            brand = ""
            size = ""
            link = ""
            model_from_page = ""
            amazon_title = ""
            amazon_variations = {}

            if scraper:
                try:
                    scraped = scraper.get_product_info(asin=asin, fallback_query=product_name)
                except Exception as e:
                    logger.warning(f"Scraping failed for ASIN={asin}: {e}")
                    scraped = {}

                brand = scraped.get("brand", "") or ""
                size = scraped.get("size", "") or ""
                link = scraped.get("link", "") or ""
                model_from_page = scraped.get("model", "") or ""
                amazon_title = scraped.get("title", "") or ""
                amazon_variations = scraped.get("variations", {}) or {}

                _sleep_human()

            # -----------------------------
            # Create row now (placeholders), compute qty/unit using best current size text
            # -----------------------------
            taglia_initial = size or ""  # LLM may improve later

            size_text_for_calc = taglia_initial or size or amazon_title or product_name
            qty_out, unita_out, unit_price = _derive_qty_unit_and_unit_price(
                base_qty=base_qty,
                total_owed=total_owed,
                size_text=size_text_for_calc,
            )

            row_out = {
                "index": len(out_rows) + 1,
                "Acquisto": (product_name.split(",")[0].strip() if product_name else ""),
                "Costo": float(total_owed) if total_owed is not None else "",
                "Data": order_day if order_day else "",
                "Categoria": "Altro",
                "Negozio": negozio,
                "Marca": brand,
                "Modello": model_from_page or "",
                "Taglia": taglia_initial,
                "Quantità": qty_out,
                "Prezzo unità": float(unit_price) if unit_price is not None else "",
                "Unità": unita_out,
                "Link": link,
                "Note": product_name,
                "Order ID": order_id,

                # Internal fields for later recompute (not exported)
                "_product_name": product_name,
                "_model_from_page": model_from_page,
                "_scraped_size": size,
                "_amazon_title": amazon_title,
                "_base_qty": base_qty,
                "_total_owed": total_owed,
            }

            out_rows.append(row_out)

            item_id = f"{order_id}::{asin or 'NOASIN'}::{i}"
            row_by_id[item_id] = row_out

            # -----------------------------
            # Batch LLM (optional) - separated queues
            # -----------------------------
            if use_llm:
                item = {
                    "id": item_id,
                    "product_name": product_name,
                    "scraped_brand": brand,
                    "scraped_model": model_from_page,
                    "scraped_size": size,
                    "amazon_title": amazon_title,
                    "variations_json": amazon_variations,
                }

                if negozio == "Amazon Fresh":
                    pending_fresh.append(item)
                    if len(pending_fresh) >= batch_size:
                        _flush_batch(pending_fresh, "Amazon Fresh")
                        pending_fresh.clear()
                else:
                    pending_amazon.append(item)
                    if len(pending_amazon) >= batch_size:
                        _flush_batch(pending_amazon, "Amazon")
                        pending_amazon.clear()

        # -----------------------------
        # Flush remaining batches
        # -----------------------------
        if use_llm and pending_amazon:
            _flush_batch(pending_amazon, "Amazon FINAL")
            pending_amazon.clear()

        if use_llm and pending_fresh:
            _flush_batch(pending_fresh, "Amazon Fresh FINAL")
            pending_fresh.clear()

    finally:
        if scraper:
            scraper.close()

    # -----------------------------
    # Sort by date ascending + reindex
    # -----------------------------
    out_rows.sort(
        key=lambda r: (r.get("Data") is None or r.get("Data") == "", r.get("Data") or datetime.max)
    )
    for idx, r in enumerate(out_rows, start=1):
        r["index"] = idx

    logger.info(f"Writing XLSX with {len(out_rows)} rows...")
    write_xlsx(out_rows, output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Amazon order CSV export into normalized XLSX.")
    parser.add_argument("--root-dir", required=True, help="Root folder containing the 'amazon' directory with dated 'YYYY-MM-DD - Your Orders' folders.")
    parser.add_argument("--output",  default="", help="Optional override output XLSX path. If empty, saved inside latest 'YYYY-MM-DD - Your Orders' folder as output.xlsx")
    parser.add_argument("--headless", action="store_true", help="Run Chrome in headless mode.")
    parser.add_argument("--no-llm", action="store_true", help="Disable DeepSeek LLM enrichment.")
    parser.add_argument("--no-scraping", action="store_true", help="Disable Amazon scraping enrichment.")
    parser.add_argument("--max-rows", type=int, default=None, help="Process only first N rows (debug).")
    parser.add_argument("--is-raspberry", action="store_true", help="Use os_environ=True like AutoScout on Raspberry/Linux.")
    parser.add_argument("--start-date", type=str, default="", help="Process orders from this date (inclusive). Formats: YYYY-MM-DD or DD/MM/YYYY.")

    args = parser.parse_args()

    input_csv, auto_output_xlsx = _resolve_amazon_csv_from_root(args.root_dir)
    output_xlsx = args.output.strip() or auto_output_xlsx
    auto_backups_dir = Path(args.root_dir) / "backup_google_sheet_shopping"
    backups_dir = str(auto_backups_dir) if auto_backups_dir.exists() and auto_backups_dir.is_dir() else ""

    logger.info(f"Root dir: {args.root_dir}")
    logger.info(f"Input CSV: {input_csv}")
    logger.info(f"Output XLSX: {output_xlsx}")
    logger.info(f"Backups dir: {backups_dir if backups_dir else '[not found]'}")

    transform(
        input_csv=input_csv,
        output_xlsx=output_xlsx,
        headless=args.headless,
        use_llm=not args.no_llm,
        use_scraping=not args.no_scraping,
        max_rows=args.max_rows,
        is_raspberry=args.is_raspberry,
        start_date=args.start_date,
        backups_dir=backups_dir
    )


if __name__ == "__main__":
    main()
